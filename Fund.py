import requests
import pandas as pd
import numpy as np
from lxml import etree
import tkinter
import re
import os
import time
from openpyxl import load_workbook
import openpyxl
from openpyxl.styles import Font,colors #导入样式的字体大小和颜色函数（注意大小写）
from pandas import ExcelWriter
import keyboard
"""
基金数据查询工具
说明:
通过输入基金代码 得到基金最新估算涨幅 近三年内涨幅 基金股票持仓情况 基金股票涨幅 基金经理姓名
功能:数据清洗 每次查询清理掉重复的数据 留下最新的数据
数据抓取:通过requests库抓取各种数据包以及源码标签内容
通过算法进行字体颜色渲染 从而一眼看出是涨还是跌
数据更新:通过线程池将所有基金的更新任务分配给多个线程 提高效率 (由于只加文件锁指针指向不明导致数据混淆 若再将锁前置 则多线程的意义不大 故取消此方法改用单线程)
"""

class spider():
      def __init__(self,code,mode):
          self.mode=mode
          self.code=code
          self.Answer1=str()
          self.headers={
              'User-Agent': 'Mozilla / 5.0(Windows NT 10.0;Win64;x64) AppleWebKit / 537.36(KHTML, likeGecko) Chrome / 92.0.4515.159Safari / 537.36Edg / 92.0.902.84'
          }
          self.IndexList=[]
          self.StartIndex = 0
          self.q=0
          self.choose=str()

      def start(self):
          try:
              self.get_html(self.code)
          except ConnectionError:  # 若短时间内发送请求过多
              print("查询过于频繁 请稍后再试")
              time.sleep(4)
              self.start()
     # def __str__(self): 在python中调用print()打印实例化对象时会调用__str__()如果__str__()中有返回值，就会打印其中的返回值。
     #     return "该只基金基金代号为"+self.code
      def get_html(self,code) -> "Source Code": #获取网页源码 从而实例化tree对象
          self.code=code
          url="http://fund.eastmoney.com/%s.html?spm=search"
          PageUrl=format(url%self.code)
          resp0 = requests.get(headers=self.headers,url=PageUrl)
          resp0.encoding="utf-8"
          self.resp1=resp0.text

      def crawl_recent_data(self):  #抓取其他数据 如近3年涨幅  股票持仓分布 基金经理  近11个交易日每日涨幅
          try:
             tree=etree.HTML(self.resp1)
             Name = tree.xpath("//*[@id='body']/div[11]/div/div/div[1]/div[1]/div/text()")[0]
             self.RisePercentage=[]
             for ind in range(2,10): #抓取近3年涨幅
                 value = tree.xpath(f'//*[@id="increaseAmount_stage"]/table/tr[2]/td[{ind}]/div/text()')
                 if len(value)!=0:
                    self.RisePercentage.append(tree.xpath(f'//*[@id="increaseAmount_stage"]/table/tr[2]/td[{ind}]/div/text()')[0])#浏览器会对html文本进行一定的规范化
                   #所以会自动在路径中加入tbody，导致读取失败，在此处直接在路径中去除tbody即可。
          except IndexError: #没有该基金
                 self.Answer1="Fault"
                 return self.Answer1
          try:
             url1 = "http://fundgz.1234567.com.cn/js/%s.js" #动态加载的数据无法通过xpath抓取 正则表达式也不行 源码中不含不在DOM树中的标签的内容 故需要分析数据包进行抓包
             DataUrl = format(url1 % self.code) #网址根据基金代码变化 故拼接
             resp0 = requests.get(headers=self.headers,url=DataUrl)  # 存有基金代码 基金名称 基金最后一天更新日期 上一次收盘时净值 涨幅 最后一次净值更新时的时间
             if resp0.status_code == 200: #判断网址是否正确 是否可以正常爬取
                 self.Answer1="NoProblem"
                 resp0 = resp0.text
                 self.resp0 = eval(resp0.replace("jsonpgz", "").replace("(", "").replace(")", "").replace(";",
                                                                                                          ""))  # 规范抓到的数据 并将数据类型从str转为dict:json.loads(resp0)也可成功转换
                 self.resp0["gszzl"]+="%" #给预估的净值加个%
                 self.listA = [self.resp0["fundcode"], self.resp0["name"], self.resp0["jzrq"], self.resp0["dwjz"],
                          self.resp0["gsz"], self.resp0["gszzl"], self.resp0["gztime"]]
                 self.choose = "ListA"
                 self.listA = self.listA + self.RisePercentage
                 return self.ShowData()
             else:
                 self.Answer1 = "Error"
                 return self.Answer1
          except SyntaxError:  #若数据包为空
                 self.RisePercentage = [self.code,Name,'无', '无', '无', '无', '无'] + self.RisePercentage
                 self.Answer1 = "Error"
                 self.choose = "RList"
                 return self.ShowData()

      def FontChange(self) -> "All the font style would be changed when finshed":
          table = openpyxl.load_workbook('基金.xlsx')  # 该库不支持csv 支持xlsx等excel文件
          sheet = table["Sheet1"]  # 选定Sheet1表单进行操作
          row = len(self.df)+1  # 获取表格行数 但最后一行的索引是行数+1 由于是拼接 无法省去渲染字体颜色的时间 每次都得重新渲染所有字体颜色
          while  row > 1:  #当字体有颜色 或遍历到列标时 停止更新字体颜色
                for col in range(6, 16):
                    if col == 7 :  #第7列为日期 无法判别 故跳过此次循环
                       continue
                    if str(sheet.cell(row=row, column=col).value) == "--" or "无":
                       continue
                    sheet.cell(row=row, column=col).value = float(str(sheet.cell(row=row, column=col).value).replace("%", ""))  # 先去掉百分号转换为不带百分号的浮点数进行正负判断
                    sheet.cell(row=row, column=col).font = Font(
                    color='FF0000' if float(sheet.cell(row=row,column=col).value) > 0.0 else '006400',
                    bold=True)  # 将基金涨幅那一列的字体版式进行逐一替换
                    sheet.cell(row=row, column=col).value = str(sheet.cell(row=row,column=col).value) + "%" #将原为浮点百分数的小数转换为百分数
                    # 设置净值涨幅颜色 赚就红 亏就绿 仅支持16进制ARGB颜色格式
                row -= 1
          table.save('基金.xlsx')  # 保存文件
          table.close()

      def ShowData(self) -> "show the data and save it in xlsx":
          NameList = ["基金代码","基金名称","收盘日","收盘净值","当前净值","净值涨幅","更新时间","近1周","近1月","近3月","近6月","今年来","近1年","近2年","近3年"]
          if self.choose == "ListA": #若抓的是普通基金
              Index = self.listA    #获取数据 +多少根据列标定
          else: #若抓的是ETF等基金
              Index = self.RisePercentage
          RealIndex = dict(zip(NameList,Index)) #将列表合成为字典
          if self.q == 0 :   #若第一次查找
             self.df = pd.DataFrame.from_dict(RealIndex,orient='index').T  #将字典对应的二维数组反置 使key为列标
             self.q += 1
             if os.path.exists("基金.xlsx") == False : #若不存在该文件
                BlockFrame = pd.DataFrame()
                BlockFrame.to_excel("基金.xlsx")
                time.sleep(2)
          elif os.path.exists("基金.xlsx") and self.q>0:  #若不是第一次查找且该文件存在
              df1 = pd.Series(RealIndex)   #初始化一维列表
              self.df = self.df.append(df1,ignore_index=True)    #添加新的基金数据
          return self.data_processing()


      def data_processing(self) -> "Process the data and integrate/concatenate them":
               data = pd.read_excel("基金.xlsx")  # 读取文件
               self.RowNum = data.shape[0]
               if self.RowNum == 0:  # 如果excel为空 不去掉列标
                   self.df.to_excel("基金.xlsx", index=False, encoding="gb2312")
                   self.FontChange()  # 修改字体颜色
               elif self.RowNum != 0:  # 若该文件已存在内容 去掉列标 但to_excel方法会覆盖旧excel 创建新的表 故需提取上次存进的所有内容 从而进行拼接
                   df2 = data.iloc[0:self.RowNum]  # 即取出上次保存的文件的每行数据
                   NewDf = [df2,self.df]  # 两个DF对象组合为新的df对象
                   self.df = pd.concat(objs=NewDf, axis=0)  # 进行纵向拼接 存入新excel表
                   self.df.drop_duplicates(subset="基金名称", keep='last',
                                           inplace=True)  # 去重操作  根据列标基金名称查找重复项 从上向下查找 由于插入时从尾部插入 故keep='last' 否则keep='first'
                   self.df.to_excel("基金.xlsx", index=False, encoding="gb2312")
                   self.FontChange()
                   if self.mode == "inquire" and self.choose != "RList":  #若是查询模式 数据处理完直接强制退出
                      return self.listA
                   else:  #若抓取的是ETF之类的基金
                      return self.RisePercentage

      def update_data(self) -> "A function which is constucted to update realtime data":
          Spider=spider(1,"update")
          data = pd.read_excel("基金.xlsx")  # 读取文件
          RowNum = data.shape[0] #取出行数
          CodeList = list(data.loc[0:RowNum, '基金代码'])  # 将基金代码放入列表
          NewCodeList = [(6 - len(str(p))) * '0' + str(p) for p in CodeList]  #给基金代码补0
          for code in NewCodeList:
              Spider.get_html(code)
              Spider.crawl_recent_data()
          print("基金数据更新完毕")

      def __del__(self):
          return 0

if __name__ == "__main__":
    np.set_printoptions(suppress=True)  # 取消科学计数法
    print("按u键进入数据更新模式\n按shift键进入数据获取模式")
    if keyboard.read_key() == "u":
        Spider = spider(1, "update")
        Spider.update_data()
    elif keyboard.read_key() == "shift":
        Code = input("请输入您要查询的基金的基金代码:")
        Spider = spider(Code, "inquire")
        Spider.start()